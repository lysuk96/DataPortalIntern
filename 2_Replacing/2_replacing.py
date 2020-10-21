import openpyxl


#엑셀 열기
def open_xl():
    global ws, wb

    wb = openpyxl.load_workbook('1.xlsx')
    ws = wb.active

# 엑셀 저장 및 종료
def save_exit_xl():
    global wb

    wb.save("1.xlsx")
    wb.close()
    exit()

def chck_str(tmp) :
    chck_arr = ["&lt;","&gt;","&#38;","&#x27;","&quot;","&#35;","&#40;","&#41;","&#x2F;"]
    for i in range(0,9) :
        if chck_arr[i] in tmp :
            return True, chck_arr[i]
    return False, None

def chg_str(tmp) :
    chck_map = {"&lt;":"<","&gt;":">","&#38;":"&","&#x27;":"\'","&quot;":"\"","&#35;":"#","&#40;":"(","&#41;":")","&#x2F;":"/"}
    while (True):
        flag, to_chg = chck_str(tmp)
        if flag is False : break
        tmp = tmp.replace(to_chg, chck_map[to_chg])
    return tmp

def fnd_row():
    global row_limit, ws

    chg_row = input("Which row do you want to change? : ")
    chg_row = chg_row.upper()
    
    row_limit=1
    while(row_limit<40) :
        if(ws.cell(1,row_limit).value==None):
            print("No such row in this sheet")
            save_exit_xl()
        elif(ws.cell(1,row_limit).value==chg_row):
            break
        row_limit=row_limit+1

def chg_all():
    global row_limit

    for i in range(2, ws.max_row + 1) :
        txt = chg_str(ws.cell(i,row_limit).value)
        ws.cell(i,row_limit).value = txt


# Main
open_xl()
fnd_row()
chg_all()
save_exit_xl()
