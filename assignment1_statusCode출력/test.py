import requests 
import openpyxl
import threading

# 엑셀파일 열기
wb = openpyxl.load_workbook('1.xlsx')
 
# 현재 Active Sheet 얻기
ws = wb.active

def doTheThing():
    row_index = r[0].row
    url = 'http://' + r[0].value
    
    try :
        threading.Timer(0.1, timer)
        response = requests.get(url) 
        flag = response.status_code

        if (flag >= 200 and flag < 300) :
            ws.cell(row=row_index, column=2).value = 'y'
        else :
            ws.cell(row=row_index, column=2).value = 'n'
            ws.cell(row=row_index, column=3).value = flag
        
    except requests.exceptions.MissingSchema :
        print(str(row_index) + "\t\tMissingSchema error : \t\t" + url)
    except requests.exceptions.ConnectionError :
        print(str(row_index) + "\t\tConnectionError error : \t\t" + url)
    except requests.exceptions.InvalidURL :
        print(str(row_index) + "\t\tInvalidURL error : \t\t" + url)
    except Exception as e :
        print(str(row_index) + "\t\tTime error : \t\t" + url)

def timer():
    raise Exception()

for r in ws.rows:
    row_index = r[0].row
    url = r[0].value

        
    try :
        threading.Timer(0.1, timer)
        response = requests.get(url) 
        flag = response.status_code

        if (flag >= 200 and flag < 300) :
            ws.cell(row=row_index, column=2).value = 'y'
        else :
            ws.cell(row=row_index, column=2).value = 'n'
            ws.cell(row=row_index, column=3).value = flag
        
    except requests.exceptions.MissingSchema :
        doTheThing()
        continue
    except requests.exceptions.ConnectionError :
        print(str(row_index) + "\t\tConnectionError error : \t\t" + url)
        continue
    except requests.exceptions.InvalidURL :
        print(str(row_index) + "\t\tInvalidURL error : \t\t" + url)
        continue
    except Exception as e :
        print(str(row_index) + "\t\tTime error : \t\t" + url)
        continue
 
# 엑셀 파일 저장
wb.save("2.xlsx")
wb.close()
